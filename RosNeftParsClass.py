from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as exp_conds
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, ElementNotVisibleException, TimeoutException
from openpyxl import Workbook
from datetime import datetime
import sys, json, threading, time, contextlib
from PyQt5.QtWidgets import QWidget, QPushButton, QApplication, QLineEdit, QLabel, QTextEdit
from PyQt5.QtCore import QCoreApplication
from os import path

class RosNeftParse(QWidget):

    def __init__(self, parsUrl='http://zakupki.rosneft.ru/zakupki/arhive'):
        super().__init__()
        self.url = parsUrl
        self.searchFromDate = ' '
        self.searchTillDate = ' '
        self.searchKey = ' '
        self.stopFlag = False
        self.initUI()

    def onStart(self):
        self.stopFlag = False
        threading.Thread(target=self.parseSite).start()
        self.startBtn.setVisible(False)
        self.stopBtn.setVisible(True)
        self.quitBtn.setVisible(False)
        self.statusLbl.setText('Status: In Process')
        self.statusLbl.adjustSize()

    def onStop(self):
        self.stopBtn.setVisible(False)
        self.stopFlag = True
        time.sleep(5)
        self.startBtn.setVisible(True)
        self.quitBtn.setVisible(True)
        self.statusLbl.setText('Status: Finished and Ready again!')
        self.statusLbl.adjustSize()

    def initUI(self):

        self.textField = QTextEdit(self)
        self.textField.move(270, 100)

        self.statusLbl = QLabel(self)
        self.statusLbl.setText('Status: Ready')
        self.statusLbl.move(270, 300)

        self.quitBtn = QPushButton('Quit', self)
        self.quitBtn.clicked.connect(QCoreApplication.instance().quit)
        self.quitBtn.resize(self.quitBtn.sizeHint())
        self.quitBtn.move(50, 180)

        self.startBtn = QPushButton('Start', self)
        self.startBtn.clicked[bool].connect(self.onStart)
        self.startBtn.resize(self.startBtn.sizeHint())
        self.startBtn.move(50, 120)

        self.stopBtn = QPushButton('Stop', self)
        self.stopBtn.clicked.connect(self.onStop)
        self.stopBtn.resize(self.quitBtn.sizeHint())
        self.stopBtn.move(50, 150)
        self.stopBtn.setVisible(False)

        self.fromLbl = QLabel(self)
        self.fromLbl.setText('From Date\nYYYY-MM-DD')
        self.fromLbl.move(150, 75)

        self.tillLbl = QLabel(self)
        self.tillLbl.setText('Till Date\nYYYY-MM-DD')
        self.tillLbl.move(150, 160)

        self.keyLbl = QLabel(self)
        self.keyLbl.setText('Key Words')
        self.keyLbl.move(310, 50)

        self.fromDateLine = QLineEdit(self)
        self.fromDateLine.setGeometry(150, 50, 100, 20)
        self.tillDateLine = QLineEdit(self)
        self.tillDateLine.setGeometry(150, 140, 100, 20)
        self.fromDateLine.textChanged[str].connect(self.onChangedFrom)
        self.tillDateLine.textChanged[str].connect(self.onChangedTill)
        self.keyWordsLine = QLineEdit(self)
        self.keyWordsLine.setGeometry(380, 50, 150, 20)
        self.keyWordsLine.textChanged[str].connect(self.onChangedKey)
        self.setGeometry(500, 250, 550, 350)
        self.setWindowTitle('RosNeftParser')
        self.show()

    def onChangedFrom(self, text):
        self.searchFromDate = text

    def onChangedTill(self, text):
        self.searchTillDate = text

    def onChangedKey(self, text):
        self.searchKey = text

    def parseLink(self, link):
        self.subDriver.get(link)
        result = {}
        while True:
            try:
                title = WebDriverWait(self.subDriver, 10).until(exp_conds.presence_of_element_located(
                    (By.CLASS_NAME, 'title')))
                break
            except TimeoutException:
                print('Waiting for page loaded...')

        result['Номер закупки'] = self.subDriver.find_element_by_xpath(
            '//*[@id="main"]/table/tbody/tr[2]/td/div/strong[1]').text
        result['Дата публикации'] = self.subDriver.find_element_by_xpath(
            '//*[@id="main"]/table/tbody/tr[2]/td/div/strong[2]').text
        passFlag = False
        try:
            tempdata = self.subDriver.find_element_by_xpath(
                        '//*[@id="main"]/table/tbody/tr[2]/td/div/strong[3]').text + ' по '+ self.subDriver.find_element_by_xpath(
                '//*[@id="main"]/table/tbody/tr[2]/td/div/strong[4]').text
            if 'Способ' not in tempdata:
                result['Срок подачи заявок'] = tempdata
            else:
                result['Способ закупки'] = tempdata.split(' по ')[0].split(' - ')[1]
                result['Статус закупки'] = tempdata.split(' по ')[1].split(' - ')[1]
                passFlag = True
        except NoSuchElementException:
            pass
        if not passFlag:
            try:
                result['Способ закупки'] = self.subDriver.find_element_by_xpath(
                            '//*[@id="main"]/table/tbody/tr[2]/td/div/strong[5]').text.split(' - ')[1]
            except NoSuchElementException:
                pass
            try:
                result['Статус закупки'] = self.subDriver.find_element_by_xpath(
                            '//*[@id="main"]/table/tbody/tr[2]/td/div/strong[6]').text.split(' - ')[1]
            except NoSuchElementException:
                pass
        result['Наименование закупки'] = title.text
        result['Ссылка на закупку'] = link
        tables = self.subDriver.find_elements_by_xpath('//*[@id="main"]/table/tbody/tr[2]/td/table')
        for tableID in range(1, len(tables) + 1):
            tableRows = self.subDriver.find_elements_by_xpath(
                '//*[@id="main"]/table/tbody/tr[2]/td/table[' + str(tableID) + ']/tbody/tr')
            for rowID in range(1, len(tableRows) + 1):
                rowCells = self.subDriver.find_elements_by_xpath('//*[@id="main"]/table/tbody/tr[2]/td/table['
                                                         + str(tableID) + ']/tbody/tr[' + str(rowID) + ']/td')
                key = rowCells[0].text
                value = rowCells[1].text
                if 'Адрес' in value:
                    key = 'Адрес'
                    value = self.subDriver.find_element_by_class_name('contact-adress').text.split('\n')[1]

                elif 'Организатор' in key:
                    value = value.split('\n')[0]
                elif 'Извещение о закупке' in key or 'Документация' in key or 'Протоколы' in key:
                    elem = self.subDriver.find_element_by_xpath('//*[@id="main"]/table/tbody/tr[2]/td/table['
                                                        + str(tableID) + ']/tbody/tr['
                                                        + str(rowID) + ']/td[2]/div/a')
                    value = elem.get_attribute('href')
                result[key] = value
        return result

    def parseSite(self):
        self.chromeDriverPath = path.abspath('chromedriver')
        self.driver = webdriver.Chrome(self.chromeDriverPath)
        self.driver.get(self.url)
        self.subDriver = webdriver.Chrome(self.chromeDriverPath)
        self.parsedData = {}
        self.currDate = datetime.now().strftime("_%d-%B-%Y_%I-%M%p_")
        self.fields = ['Ссылка на закупку', 'Номер закупки', "Дата публикации", 'Срок подачи заявок',
                       'Наименование закупки', 'Статус закупки', 'Способ закупки', 'Организатор', 'Адрес',
                  "Сведения о начальной (максимальной) цене договора (цене лота)", 'Общий классификатор закупки',
                       "Требования к участникам"]
        self.ROWID = 2

        searchBtn = WebDriverWait(self.driver, 10).until(exp_conds.presence_of_element_located(
            (By.XPATH, '//*[@id="views-exposed-form-zakupki-page-5"]/div/fieldset/legend/a[2]')))

        try:
            if not self.driver.find_element_by_xpath('//*[@id="edit-field-zakup-begin-value-datepicker-popup-0"]').is_displayed():
                searchBtn.click()
        except NoSuchElementException:
            searchBtn.click()
        except ElementNotVisibleException:
            searchBtn.click()

        while True:
            try:
                WebDriverWait(self.driver, 10).until(exp_conds.presence_of_element_located(
                    (By.XPATH, '//*[@id="edit-field-zakup-begin-value-datepicker-popup-0"]'))).send_keys(self.searchFromDate)
                WebDriverWait(self.driver, 10).until(exp_conds.presence_of_element_located(
                    (By.XPATH, '//*[@id="edit-field-zakup-end-value-datepicker-popup-0"]'))).send_keys(self.searchTillDate)
                WebDriverWait(self.driver, 10).until(exp_conds.presence_of_element_located(
                    (By.XPATH, '//*[@id="edit-title"]'))).send_keys(self.searchKey)
                WebDriverWait(self.driver, 10).until(exp_conds.presence_of_element_located((By.ID, 'edit-submit-zakupki'))).click()

                WebDriverWait(self.driver, 10).until(exp_conds.presence_of_element_located(
                    (By.XPATH, '//*[@id="main"]/table/tbody/tr[2]/td/div/div[3]/table[2]/thead/tr/th[1]/a')))
                break
            except ElementNotVisibleException:
                searchBtn.click()

        wb = Workbook()
        exlFile = path.abspath('rosneft' + self.currDate + '(' + self.searchFromDate + '_' + self.searchTillDate + ').xlsx')
        jsonFile = path.abspath('data_' + self.currDate + '.json')
        nextPageUrl = ''

        with open(jsonFile, 'a') as file:
            file.write('[')

        wb.save(exlFile)
        sheet1 = wb.active
        sheet1.title = 'RosNeft'
        sheet1['A1'] = '№'
        for col in range(2, len(self.fields) + 2):
            _ = sheet1.cell(column=col, row=1, value=self.fields[col - 2])

        while nextPageUrl is not None:
            while True:
                try:
                    WebDriverWait(self.driver, 10).until(exp_conds.presence_of_element_located(
                        (By.XPATH, '//*[@id="main"]/table/tbody/tr[2]/td/div/div[3]/table[2]/thead/tr/th[1]/a')))
                    break
                except TimeoutException:
                    print('Waiting for page loaded...')

            nextPageUrl = self.driver.find_element_by_xpath('//*[@id="main"]/table/tbody/tr[2]/td/div/div[4]/ul/li['
                                                       + str(
                len(self.driver.find_elements_by_xpath('//*[@id="main"]/table/tbody/tr[2]/td/div/div[4]/ul/li')) - 1)
                                                       + ']/a')
            if nextPageUrl is None or 'следующая' not in nextPageUrl.text:
                nextPageUrl = None
            else:
                nextPageUrl = nextPageUrl.get_attribute('href')
            links = []
            for i in range(1, len(self.driver.find_elements_by_xpath(
                    '//*[@id="main"]/table/tbody/tr[2]/td/div/div[3]/table[2]/tbody/tr')) + 1):
                links.append(
                    self.driver.find_element_by_xpath('//*[@id="main"]/table/tbody/tr[2]/td/div/div[3]/table[2]/tbody/tr['
                                                 + str(i) + ']/td[1]/a').get_attribute('href'))
            links = set(links)
            for page in links:
                if self.stopFlag is True:
                    self.subDriver.close()
                    self.driver.close()
                    return
                result = self.parseLink(page)
                if result is None:
                    continue
                with open(jsonFile, 'a') as jsonfile:
                    json.dump(result, jsonfile, ensure_ascii=False, indent=4)
                    jsonfile.write(',')
                _ = sheet1.cell(column=1, row=self.ROWID, value=str(self.ROWID - 1))
                for col in range(2, len(self.fields) + 2):
                    _ = sheet1.cell(column=col, row=self.ROWID, value=str(result.get(self.fields[col - 2])))
                self.ROWID += 1
                wb.save(exlFile)
                self.textField.append(json.dumps(result, indent=4, ensure_ascii=False))
            if nextPageUrl is None:
                self.onStop()
                break
            else:
                self.driver.get(nextPageUrl)
        self.subDriver.close()
        self.driver.close()
        with open(jsonFile, 'a') as file:
            file.write(']')
        return

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = RosNeftParse()
    sys.exit(app.exec_())
